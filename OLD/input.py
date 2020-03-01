from openpyxl import load_workbook
import settings
classes = []
book = load_workbook(settings.uch_plan_sheet_name, data_only=True)
page = book[settings.uch_plan_page_name]
def create_dict_mas (classes, dict, val=''):
    '''Создание словаря с заголовками из массива со значениями val'''
    a={}
    for i in classes:
        if val=='':
            a[i]={}
        else:
            a[i]={}
        dict.update(a)
        a={}


'''ОБРАБОТКА УЧЕБНОГО ПЛАНА'''
print('НАЧАЛО ОБРАБОТКИ УЧЕБНОГО ПЛАНА')

'''Записываем в переменную classes все классы'''
for i in range(int(settings.uch_plan_class_column_start),int(settings.uch_plan_class_column_stop)+1):
    a = page.cell(row = int(settings.uch_plan_class_row), column = i).value
    classes.append(a[0])

class_subject = {}
create_dict_mas(classes,class_subject)


for r in range(int(settings.uch_plan_subj_row_start), int(settings.uch_plan_subj_row_stop)+1):
    cur_subj = page.cell(row = r, column = int(settings.uch_plan_subj_column)).value
    for c in range(int(settings.uch_plan_class_column_start),int(settings.uch_plan_class_column_stop)+1):
        cur_class = page.cell(row=int(settings.uch_plan_class_row), column=c).value[0]

        if type(page.cell(row = r, column = c).value) == int:
            class_subject[str(cur_class)][cur_subj] = page.cell(row = r, column = c).value

print('КОНЕЦ ОБРАБОТКИ УЧЕБНОГО ПЛАНА')

'''КОНЕЦ ОБРАБОТКИ УЧЕБНОГО ПЛАНА'''


'''ОБРАБОТКА РАСПИСАНИЯ'''
rasp_sheet = load_workbook(settings.rasp_sheet_name)
proizv_cal_page = rasp_sheet[settings.rasp_proizv_cal_page_name]
proizv_cal={}
'''ОБРАБОТКА ПЕРВОЙ СТРОКИ МЕСЯЦЕВ'''
row = [str(s) for s in proizv_cal_page.iter_rows(max_row=int(settings.rasp_proizv_cal_row1), values_only=True)]
row=row[0]
row = [str(s) for s in row.split(',')]
for i in row:

    if i==' None' or i==' None)' or i=='(None':
        continue
    else:
        format=''
        for s in i:
            if s.isdigit() or s.isalpha() or s==' ':
                format+=s
            else:
                continue
        format = [str(s) for s in format.split()]
        proizv_cal[format[0]]=int(format[1])
'''ОБРАБОТКА ВТОРОЙ СТРОКИ МЕСЯЦЕВ'''
row = [str(s) for s in proizv_cal_page.iter_rows(min_row=int(settings.rasp_proizv_cal_row2), max_row=int(settings.rasp_proizv_cal_row2)+1, values_only=True)]
row=row[0]
row = [str(s) for s in row.split(',')]
for i in row:

    if i==' None' or i==' None)' or i=='(None':
        continue
    else:
        format=''
        for s in i:
            if s.isdigit() or s.isalpha() or s==' ':
                format+=s
            else:
                continue
        format = [str(s) for s in format.split()]
        proizv_cal[format[0]]=int(format[1])
'''ОБРАБОТКА ТРЕТЬЕЙ СТРОКИ МЕСЯЦЕВ'''
row = [str(s) for s in proizv_cal_page.iter_rows(min_row=int(settings.rasp_proizv_cal_row3), max_row=int(settings.rasp_proizv_cal_row3)+1, values_only=True)]
row=row[0]
row = [str(s) for s in row.split(',')]
for i in row:

    if i==' None' or i==' None)'  or i=='(None':
        continue
    else:
        format=''
        for s in i:
            if s.isdigit() or s.isalpha() or s==' ':
                format+=s
            else:
                continue
        format = [str(s) for s in format.split()]
        proizv_cal[format[0]]=int(format[1])
'''КОНЕЦ ОБРАБОТКИ ПРОИЗВОДСТВЕННОГО КАЛЕНДАРЯ'''

module_plan_page=rasp_sheet[settings.module_plan_page_name]
'''СЧИТЫВАНИЕ НАЗВАНИЙ КЛАССОВ'''
for i in module_plan_page(row=)